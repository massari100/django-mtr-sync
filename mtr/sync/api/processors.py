from __future__ import unicode_literals

import os

from django.utils.translation import gettext_lazy as _
from django.utils.six.moves import range
from django.utils import timezone

from .manager import manager
from .signals import export_started, export_completed
from ..settings import LIMIT_PREVIEW, FILE_PATH


class Processor(object):
    """Base implementation of import and export operations"""

    position = 0
    file_format = None
    file_description = None

    def __init__(self, settings):
        self.settings = settings
        self.report = None

    def prepare(self):
        pass

    def col(self, value):
        """Return column index for given name"""

        raise NotImplementedError

    def write(self, row, value, cells=None):
        """Independend write to cell method"""

        raise NotImplementedError

    def read(self, row, value, cells=None):
        """Independend read from cell method"""

        raise NotImplementedError

    def set_dimensions(
            self, start_row, start_col, end_row, end_col, preview=False):
        """Return start, end table dimensions"""

        start = {'row': start_row, 'col': start_col}
        end = {'row': end_row, 'col': end_col}

        if self.settings.start_row and \
                self.settings.start_row >= start_row:
            start['row'] = self.settings.start_row - 1
            end['row'] += start['row']

        if self.settings.end_row and \
                self.settings.end_row <= end['row']:
            end['row'] = self.settings.end_row

        limit = LIMIT_PREVIEW()
        if preview and limit <= end_row:
            end_row = limit

        if self.settings.start_col:
            start_col_index = self.col(self.settings.start_col)

            if start_col_index >= start_col:
                start['col'] = start_col_index - 1
                end['col'] += start['col']

        self.start, self.end = start, end
        self.cells = range(start['col'], end['col'])
        self.rows = range(start['row'], end['row'])

        return (start, end)

    def export_data(self, data):
        """Export data from queryset to file and return path"""

        # send signal to create report
        for response in export_started.send(self.__class__, processor=self):
            self.report = response[1]

        self.set_dimensions(0, 0, data['rows'], data['cols'])

        # create export file for write
        self.create()

        # write header
        if self.settings.include_header and self.settings.fields.all():
            header_data = map(lambda f: f.name, self.settings.fields.all())

            self.write(self.start['row'], header_data)

            self.start['row'] += 1
            if self.settings.end_row:
                self.end['row'] += 1

            self.rows = range(self.start['row'], self.end['row'])

        # write data
        data = data['items']
        for row in self.rows:
            row_data = []

            for col in self.cells:
                row_data.append(next(data))

            self.write(row, row_data)

        # save external file and report
        path = FILE_PATH()(self.report, '')
        if not os.path.exists(path):
            os.makedirs(path)

        filename = '{}{}'.format(
            self.settings.filename or str(self.report.id), self.file_format)

        path = os.path.join(path, filename)
        self.save(path)

        if self.settings.id:
            self.report.settings = self.settings

        # send signal to save report
        for response in export_completed.send(
                self.__class__, report=self.report,
                date=timezone.now(),
                path=FILE_PATH()(self.report, filename, relative=True)):
            self.report = response[1]

        return self.report

    # def import_data(self):
    #     """Import data to model and return errors if exists"""

    #     raise NotImplementedError

# TODO: default import configuration, import dependencies

try:
    import xlrd
    import xlwt
except ImportError:
    pass


@manager.register
class XlsProcessor(Processor):
    file_format = '.xls'
    file_description = _('mtr.sync:Microsoft Excel 97/2000/XP/2003')

    def col(self, value):
        """Small xlrd hack to get column index"""

        if value.isdigit():
            return int(value)

        index = 0
        value = value.strip().upper()
        while index:
            if xlrd.colname(index) == value:
                return index

    def create(self):
        self._workbook = xlwt.Workbook('utf-8')
        self._worksheet = self._workbook.add_sheet(self.settings.worksheet)

    def open(self):
        self._workbook = xlrd.open_workbook(self.settings.buffer_file)
        self._worksheet = self._workbook(self.settings.worksheet)

    def write(self, row, value):
        for index, cell in enumerate(self.cells):
            self._worksheet.write(row, cell, value[index])

    def read(self, row):
        for index, cell in enumerate(self.cells):
            yield self._worksheet.cell_value(row, cell)

    def save(self, name):
        self._workbook.save(name)


try:
    import openpyxl
except ImportError:
    pass


@manager.register
class XlsxProcessor(Processor):
    file_format = '.xlsx'
    file_description = _('mtr.sync:Microsoft Excel 2007/2010/2013 XML')

    def col(self, value):
        if value.isdigit():
            return int(value)

        return openpyxl.cell.column_index_from_string(value)

    def create(self):
        self._workbook = openpyxl.Workbook(optimized_write=True)
        self._worksheet = self._workbook.create_sheet()
        self._worksheet.name = self.settings.worksheet
        self._prepend = None

        # prepend rows and cols
        if self.start['row'] > 0:
            for i in range(0, self.start['row']):
                self._worksheet.append([])

        if self.start['col'] > 0:
            self._prepend = [None, ]
            self._prepend *= self.start['col']

    def open(self):
        self._workbook = xlrd.open_workbook(self.settings.buffer_file)
        self._worksheet = self._workbook(self.settings.worksheet)

    def write(self, row, value):
        # if we have cells we need to copy from templated file all row
        # and merge cells into

        if self._prepend:
            value = self._prepend + value

        self._worksheet.append(value[:self.end['col']])

    def read(self, row, cells):
        # using optimized reader
        # for index, cell in enumerate(cells):
        #     yield self._worksheet.cell_value(row, cell)
        pass

    def save(self, name):
        self._workbook.save(name)

# import odf


# class OdsProcessor(Processor):
#     file_format = '.ods'
#     file_description = 'mtr.sync:ODF Spreadsheet'

#     def col(self, value):
#         """Small xlrd hack to get column index"""

#         if value.isdigit():
#             return int(value)

#         index = 0
#         value = value.strip().upper()
#         while index:
#             if xlrd.colname(index) == value:
#                 return index

#     def create(self):
#         self._workbook = odf.opendocument.OpenDocumentSpreadsheet()
#         self._table = odf.table.Table()
#         self._worksheet = self._workbook.spreadsheet.addElement(self._table)

#     def open(self):
#         self._workbook = xlrd.open_workbook(self.settings.buffer_file)
#         self._worksheet = self._workbook(self.settings.worksheet)

#     def write(self, row, value):
#         table_row = odf.table.TableRow()
#         self._table.addElement(table_row)
#         for index, cell in enumerate(self.cells):
#             table_cell = odf.table.TableCell()
#             table_row =
#             self._worksheet.write(row, cell, value[index])

#     def read(self, row):
#         for index, cell in enumerate(self.cells):
#             yield self._worksheet.cell_value(row, cell)

#     def save(self, name):
#         self._workbook.save(name)
