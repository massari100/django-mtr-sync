import os

os.environ['OPENPYXL_LXML'] = 'False'

import openpyxl

from ..processor import Processor
from ..manager import manager
from ...helpers import gettext_lazy as _


@manager.register('processor')
class XlsxProcessor(Processor):
    file_format = '.xlsx'
    file_description = _('Microsoft Excel 2007/2010/2013 XML')

    def create(self, path):
        self.path = path
        self.workbook = openpyxl.Workbook(optimized_write=True)
        self.worksheet = self.workbook.create_sheet()

        if self.settings:
            self.worksheet.title = self.settings.worksheet

    def open(self, path):
        # TODO: make write and read params

        self.workbook = openpyxl.load_workbook(
            path, guess_types=False, use_iterators=True,
            keep_vba=False, data_only=True, read_only=True)

        if self.settings:
            if not self.settings.worksheet:
                self.settings.worksheet = self.workbook.get_sheet_names()[0]

            self.worksheet = self.workbook.get_sheet_by_name(
                self.settings.worksheet)

            # self._read_from_start = False
            self._rows = self._worksheet.iter_rows()

            self.nrows = self.worksheet.get_highest_row()
            self.ncells = self.worksheet.get_highest_column()

        return self.workbook

    def write(self, row, value):
        self.worksheet.append(value)

    def _get_row(self, row):
        value = None
        row += 1

        try:
            if self._read_from_start:
                self._rows = self._worksheet.iter_rows()
                self._rows_counter = 0

            while self._rows_counter < row:
                self._rows_counter += 1
                value = next(self._rows)
        except StopIteration:
            return [''] * (self._max_cells + 1)

        return list(map(lambda v: v.value, value))

    def read(self, row, cells=None):
        readed = []
        value = self._get_row(row)
        cells = cells or self.cells

        for index in cells:
            try:
                readed.append(value[index])
            except IndexError:
                readed.append('')

        return readed

    def save(self, path=None):
        path = path or self.path

        self.workbook.save(path)
