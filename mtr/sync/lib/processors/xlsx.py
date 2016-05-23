import os

os.environ['OPENPYXL_LXML'] = 'False'

import openpyxl

from ..processor import Processor
from ..manager import manager
from ...translation import gettext_lazy as _


@manager.register('processor')
class XlsxProcessor(Processor):
    file_format = '.xlsx'
    file_description = _('Microsoft Excel 2007/2010/2013 XML')

    def create(self, path):
        self._path = path
        self._workbook = openpyxl.Workbook(optimized_write=True)
        self._worksheet = self._workbook.create_sheet()
        self._worksheet.title = self.settings.worksheet

        if self.start['row'] > 1:
            for i in range(0, self.start['row']):
                self._worksheet.append([])

    def open(self, path):
        self._workbook = openpyxl.load_workbook(path, use_iterators=True)

        if not self.settings.worksheet:
            self.settings.worksheet = self._workbook.get_sheet_names()[0]

        self._worksheet = self._workbook.get_sheet_by_name(
            self.settings.worksheet)

        self._read_from_start = False
        self._worksheet.calculate_dimension()
        self._rows = self._worksheet.iter_rows()
        self._rows_counter = 0
        self._max_cells = self._worksheet.get_highest_column()

        return (
            self._worksheet.get_highest_row(),
            self._max_cells)

    def write(self, row, value):
        self._worksheet.append(value)

    def _get_row(self, row):
        value = None
        row += 1

        try:
            if self._read_from_start:
                self._rows = self._worksheet.iter_rows()
                self._rows_counter = 0

            while self._rows_counter < row:
                self._rows_counter += 1
                value = next(self._rows)
        except StopIteration:
            return [''] * (self._max_cells + 1)

        return [v.value for v in value]

    def read(self, row, cells=None):
        readed = []
        value = self._get_row(row)
        cells = cells or self.cells

        for index in cells:
            try:
                readed.append(value[index])
            except IndexError:
                readed.append('')

        return readed

    def save(self):
        self._workbook.save(self._path)
